#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
首靠船风险评估表生成脚本（东莞辖区）
运行后生成：首靠船风险评估表.xlsx
"""

import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule, DataBarRule
    from openpyxl.chart import RadarChart, Reference
except ImportError:
    print("正在安装 openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule, DataBarRule
    from openpyxl.chart import RadarChart, Reference


# ============ 样式定义 ============
def create_styles(wb):
    styles = {}

    # 标题样式
    title_style = NamedStyle(name="title_style")
    title_style.font = Font(name="微软雅黑", size=16, bold=True, color="FFFFFF")
    title_style.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    title_style.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(title_style)
    styles["title"] = title_style

    # 维度标题
    dim_style = NamedStyle(name="dim_style")
    dim_style.font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    dim_style.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    dim_style.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(dim_style)
    styles["dim"] = dim_style

    # 标签
    label_style = NamedStyle(name="label_style")
    label_style.font = Font(name="微软雅黑", size=10)
    label_style.alignment = Alignment(horizontal="left", vertical="center")
    wb.add_named_style(label_style)
    styles["label"] = label_style

    # 输入单元格
    input_style = NamedStyle(name="input_style")
    input_style.font = Font(name="微软雅黑", size=10, bold=True, color="000080")
    input_style.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    input_style.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(input_style)
    styles["input"] = input_style

    # 分值（自动计算）
    score_style = NamedStyle(name="score_style")
    score_style.font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    score_style.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    score_style.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(score_style)
    styles["score"] = score_style

    # 结果
    result_style = NamedStyle(name="result_style")
    result_style.font = Font(name="微软雅黑", size=12, bold=True)
    result_style.alignment = Alignment(horizontal="center", vertical="center")
    wb.add_named_style(result_style)
    styles["result"] = result_style

    # 表头
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    header_style.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wb.add_named_style(header_style)
    styles["header"] = header_style

    return styles


# ============ 创建评估表 ============
def create_eval_sheet(wb, styles):
    ws = wb.active
    ws.title = "评估表"
    ws.sheet_properties.tabColor = "4472C4"

    # 设置列宽
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 40

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row = 1
    # 标题
    ws.merge_cells("B1:E1")
    ws["B1"] = "首靠船风险评估表（东莞辖区）"
    ws["B1"].style = "title_style"
    ws.row_dimensions[1].height = 36
    for col in range(2, 6):
        ws.cell(row=1, column=col).border = thin_border
    row += 1

    # 基本信息
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "一、基本信息"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    basic_info = [
        ("船名", "例：和谐号"),
        ("IMO编号", "例：1234567"),
        ("到港日期", "例：2026-05-10"),
        ("拟靠码头", "例：沙田港二期"),
        ("评估人", ""),
        ("评估时间", ""),
    ]
    for label, example in basic_info:
        ws[f"B{row}"] = label
        ws[f"B{row}"].style = "label_style"
        ws[f"C{row}"] = example
        ws[f"C{row}"].style = "input_style"
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    row += 1  # 空行

    # ========== 维度1：船舶固有风险 ==========
    r_dim1_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "二、船舶固有风险（本项满分30分，得分越高风险越大）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # 船龄
    ws[f"B{row}"] = "1. 船龄"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_age = DataValidation(type="list", formula1='"≤10年,10-15年,15-20年,20-25年,>25年"', allow_blank=False)
    dv_age.error = "请从下拉列表中选择"
    dv_age.errorTitle = "输入错误"
    ws.add_data_validation(dv_age)
    dv_age.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="≤10年",0,IF(C{row}="10-15年",5,IF(C{row}="15-20年",10,IF(C{row}="20-25年",15,IF(C{row}=">25年",20,0)))))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "≤10年:0分 | 10-15年:5分 | 15-20年:10分 | 20-25年:15分 | >25年:20分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_age = row
    row += 1

    # 船型
    ws[f"B{row}"] = "2. 船型"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_type = DataValidation(type="list", formula1='"普通货船/集装箱船,散货船,油轮,化学品船,液化气船(LNG/LPG),客船/滚装船"', allow_blank=False)
    ws.add_data_validation(dv_type)
    dv_type.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="普通货船/集装箱船",0,IF(C{row}="散货船",3,IF(OR(C{row}="油轮",C{row}="化学品船",C{row}="液化气船(LNG/LPG)"),5,IF(C{row}="客船/滚装船",4,0))))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "普通货船:0分 | 散货船:3分 | 油轮/化学品/LNG:5分 | 客船:4分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_type = row
    row += 1

    # 船旗国
    ws[f"B{row}"] = "3. 船旗国风险（Tokyo MOU名单）"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_flag = DataValidation(type="list", formula1='"白名单,灰名单,黑名单,未知/未列入"', allow_blank=False)
    ws.add_data_validation(dv_flag)
    dv_flag.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="白名单",0,IF(C{row}="灰名单",5,IF(C{row}="黑名单",10,3)))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "白名单:0分 | 灰名单:5分 | 黑名单:10分 | 未知:3分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_flag = row
    row += 1

    # 小计
    ws[f"B{row}"] = "小计"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"] = f"=SUM(D{r_age}:D{r_flag})"
    ws[f"C{row}"].style = "score_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    ws[f"E{row}"] = "满分30分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, italic=True, color="999999")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_dim1_end = row
    row += 1

    row += 1  # 空行

    # ========== 维度2：历史安全绩效 ==========
    r_dim2_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "三、历史安全绩效（本项满分40分，得分越高风险越大）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # PSC检查次数
    ws[f"B{row}"] = "4. 近12个月PSC检查次数"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_psc = DataValidation(type="list", formula1='"0次,1次,2次,3次及以上"', allow_blank=False)
    ws.add_data_validation(dv_psc)
    dv_psc.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="0次",0,IF(C{row}="1次",3,IF(C{row}="2次",8,15)))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "0次:0分 | 1次:3分 | 2次:8分 | ≥3次:15分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_psc = row
    row += 1

    # 滞留次数
    ws[f"B{row}"] = "5. 近36个月滞留次数"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_det = DataValidation(type="list", formula1='"0次,1次,2次,3次及以上"', allow_blank=False)
    ws.add_data_validation(dv_det)
    dv_det.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="0次",0,IF(C{row}="1次",10,IF(C{row}="2次",20,30)))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "0次:0分 | 1次:10分 | 2次:20分 | ≥3次:30分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_det = row
    row += 1

    # 缺陷数量
    ws[f"B{row}"] = "6. 近12个月PSC缺陷总数"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_def = DataValidation(type="list", formula1='"0个,1-3个,4-6个,7个及以上"', allow_blank=False)
    ws.add_data_validation(dv_def)
    dv_def.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="0个",0,IF(C{row}="1-3个",5,IF(C{row}="4-6个",10,15)))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "0个:0分 | 1-3个:5分 | 4-6个:10分 | ≥7个:15分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_def = row
    row += 1

    # 东莞历史
    ws[f"B{row}"] = "7. 东莞辖区历史检查记录"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_dg = DataValidation(type="list", formula1='"无记录,有缺陷记录,曾滞留"', allow_blank=False)
    ws.add_data_validation(dv_dg)
    dv_dg.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="无记录",0,IF(C{row}="有缺陷记录",5,15))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "无记录:0分 | 有缺陷:5分 | 曾滞留:15分（本局自建库）"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_dg = row
    row += 1

    # 小计
    ws[f"B{row}"] = "小计"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"] = f"=SUM(D{r_psc}:D{r_dg})"
    ws[f"C{row}"].style = "score_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    ws[f"E{row}"] = "满分40分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, italic=True, color="999999")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_dim2_end = row
    row += 1

    row += 1  # 空行

    # ========== 维度3：船员与配员风险 ==========
    r_dim3_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "四、船员与配员风险（本项满分25分，得分越高风险越大）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # 配员符合率
    ws[f"B{row}"] = "8. 实际配员/最低安全配员"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_crew = DataValidation(type="list", formula1='"100%,90%-99%,80%-89%,70%-79%,<70%"', allow_blank=False)
    ws.add_data_validation(dv_crew)
    dv_crew.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="100%",0,IF(C{row}="90%-99%",5,IF(C{row}="80%-89%",10,IF(C{row}="70%-79%",15,25))))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "100%:0分 | 90-99%:5分 | 80-89%:10分 | 70-79%:15分 | <70%:25分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_crew = row
    row += 1

    # 证书过期
    ws[f"B{row}"] = "9. 船员证书过期情况"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_cert = DataValidation(type="list", formula1='"无过期,1-2人过期,3人及以上过期"', allow_blank=False)
    ws.add_data_validation(dv_cert)
    dv_cert.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="无过期",0,IF(C{row}="1-2人过期",5,12))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "无过期:0分 | 1-2人:5分 | ≥3人:12分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_cert = row
    row += 1

    # 关键岗位资历
    ws[f"B{row}"] = "10. 船长/大副/轮机长在船任职时间"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_exp = DataValidation(type="list", formula1='"≥12个月,6-12个月,3-6个月,<3个月,本航次新上船"', allow_blank=False)
    ws.add_data_validation(dv_exp)
    dv_exp.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="≥12个月",0,IF(C{row}="6-12个月",2,IF(C{row}="3-6个月",5,IF(C{row}="<3个月",8,10))))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "≥12月:0分 | 6-12月:2分 | 3-6月:5分 | <3月:8分 | 新上船:10分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_exp = row
    row += 1

    # 小计
    ws[f"B{row}"] = "小计"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"] = f"=SUM(D{r_crew}:D{r_exp})"
    ws[f"C{row}"].style = "score_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    ws[f"E{row}"] = "满分25分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, italic=True, color="999999")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_dim3_end = row
    row += 1

    row += 1  # 空行

    # ========== 维度5：航行与通航风险 ==========
    r_dim5_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "五、航行与通航风险（本项满分25分，得分越高风险越大）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # 吃水余量
    ws[f"B{row}"] = "11. 吃水余量（航道水深 - 船舶吃水）"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_draft = DataValidation(type="list", formula1='"≥2.0米,1.0-2.0米,0.5-1.0米,0-0.5米,<0米(超吃水)"', allow_blank=False)
    ws.add_data_validation(dv_draft)
    dv_draft.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="≥2.0米",0,IF(C{row}="1.0-2.0米",5,IF(C{row}="0.5-1.0米",10,IF(C{row}="0-0.5米",20,25))))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "≥2m:0分 | 1-2m:5分 | 0.5-1m:10分 | 0-0.5m:20分 | <0m:25分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_draft = row
    row += 1

    # 到港时段
    ws[f"B{row}"] = "12. 预计到港时段"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_time = DataValidation(type="list", formula1='"白天(日出-日落),夜间,凌晨(00:00-06:00)"', allow_blank=False)
    ws.add_data_validation(dv_time)
    dv_time.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="白天(日出-日落)",0,IF(C{row}="夜间",3,5))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "白天:0分 | 夜间:3分 | 凌晨:5分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_time = row
    row += 1

    # 气象条件
    ws[f"B{row}"] = "13. 到港前后24小时气象预报"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_wx = DataValidation(type="list", formula1='"良好(晴/多云),一般(阴/小雨/轻雾),恶劣(大风/暴雨/大雾)"', allow_blank=False)
    ws.add_data_validation(dv_wx)
    dv_wx.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="良好(晴/多云)",0,IF(C{row}="一般(阴/小雨/轻雾)",5,12))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "良好:0分 | 一般:5分 | 恶劣:12分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_wx = row
    row += 1

    # 载货类型
    ws[f"B{row}"] = "14. 本航次载货类型"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_cargo = DataValidation(type="list", formula1='"普通货物/集装箱,包装危险品(IMDG),散装油类,散装化学品/液化气"', allow_blank=False)
    ws.add_data_validation(dv_cargo)
    dv_cargo.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="普通货物/集装箱",0,IF(C{row}="包装危险品(IMDG)",5,IF(OR(C{row}="散装油类",C{row}="散装化学品/液化气"),10,0)))'
    ws[f"D{row}"].style = "score_style"
    ws[f"E{row}"] = "普通货:0分 | 包装危险品:5分 | 散装油/化学品:10分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_cargo = row
    row += 1

    # 小计
    ws[f"B{row}"] = "小计"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="right", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"] = f"=SUM(D{r_draft}:D{r_cargo})"
    ws[f"C{row}"].style = "score_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
    ws[f"E{row}"] = "满分25分"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, italic=True, color="999999")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_dim5_end = row
    row += 1

    row += 1  # 空行

    # ========== 一票否决检查 ==========
    r_veto_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "六、一票否决项检查（若任一项为\"是\"，直接定为🔴高风险）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    veto_items = [
        ("15. 近6个月内是否有PSC滞留记录？", "r15"),
        ("16. 船龄>25年且为油轮/化学品船/液化气船？", "r16"),
        ("17. 配员符合率是否<70%？", "r17"),
        ("18. 吃水余量是否<0米（超吃水）？", "r18"),
        ("19. 关键船员证书过期≥3人或船长证书过期？", "r19"),
    ]
    r_veto_items = []
    for label, key in veto_items:
        ws[f"B{row}"] = label
        ws[f"B{row}"].style = "label_style"
        ws[f"C{row}"] = "否"
        ws[f"C{row}"].style = "input_style"
        dv_v = DataValidation(type="list", formula1='"否,是"', allow_blank=False)
        ws.add_data_validation(dv_v)
        dv_v.add(ws[f"C{row}"])
        ws[f"D{row}"] = f'=IF(C{row}="是",\"✓\",\"\")'
        ws[f"D{row}"].style = "score_style"
        ws[f"D{row}"].font = Font(name="微软雅黑", size=12, bold=True, color="C00000")
        ws[f"E{row}"] = "如选\"是\"，触发一票否决"
        ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="FF0000", bold=True)
        for col in range(2, 6):
            ws.cell(row=row, column=col).border = thin_border
        r_veto_items.append(row)
        row += 1
    r_veto_end = row - 1

    row += 1  # 空行

    # ========== 重点风险提示（新增，不纳入总分） ==========
    r_warn_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "七、重点风险提示（不纳入总分，供指挥中心和现场额外关注）"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # 桥梁净空余量
    ws[f"B{row}"] = "20. 桥梁净空余量（桥梁通航净高 - 船舶水面以上最大高度）"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_bridge = DataValidation(type="list", formula1='"≥3.0米,1.0-3.0米,0-1.0米,<0米(超高)"', allow_blank=False)
    ws.add_data_validation(dv_bridge)
    dv_bridge.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="≥3.0米","✓ 正常",IF(C{row}="1.0-3.0米","△ 关注",IF(C{row}="0-1.0米","⚠️ 高风险","❌ 禁止通航")))'
    ws[f"D{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"E{row}"] = "≥3m:正常 | 1-3m:关注 | 0-1m:⚠️高风险 | <0m:禁止通航"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_bridge = row
    row += 1

    # 驾驶员首航
    ws[f"B{row}"] = "21. 船长/驾驶员是否首次驾该船进东莞港"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_first = DataValidation(type="list", formula1='"否,船长首次,船长+驾驶员均首次"', allow_blank=False)
    ws.add_data_validation(dv_first)
    dv_first.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="否","✓ 熟悉",IF(C{row}="船长首次","△ 关注","⚠️ 高风险"))'
    ws[f"D{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"E{row}"] = "非首航:熟悉 | 船长首次:△关注 | 均首次:⚠️高风险"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_first = row
    row += 1

    # 引航安排
    ws[f"B{row}"] = "22. 引航安排"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_pilot = DataValidation(type="list", formula1='"无需引航,申请引航,强制引航"', allow_blank=False)
    ws.add_data_validation(dv_pilot)
    dv_pilot.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="无需引航","自航",IF(C{row}="申请引航","有引航","强制引航"))'
    ws[f"D{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"E{row}"] = "自航/申请引航/强制引航"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_pilot = row
    row += 1

    # 泊位匹配
    ws[f"B{row}"] = "23. 泊位吨级匹配情况"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = "请选择"
    ws[f"C{row}"].style = "input_style"
    dv_berth = DataValidation(type="list", formula1='"匹配,略微超配,明显超配"', allow_blank=False)
    ws.add_data_validation(dv_berth)
    dv_berth.add(ws[f"C{row}"])
    ws[f"D{row}"] = f'=IF(C{row}="匹配","✓ 正常",IF(C{row}="略微超配","△ 关注","⚠️ 高风险"))'
    ws[f"D{row}"].font = Font(name="微软雅黑", size=10, bold=True)
    ws[f"E{row}"] = "匹配:正常 | 略超:△关注 | 明显超配:⚠️高风险"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_berth = row
    row += 1

    # 重点船舶自动标记
    ws[f"B{row}"] = "24. 重点船舶类型（根据上方船型自动判断）"
    ws[f"B{row}"].style = "label_style"
    ws[f"C{row}"] = f'=IF(OR(C{r_type}="油轮",C{r_type}="化学品船",C{r_type}="液化气船(LNG/LPG)",C{r_type}="客船/滚装船"),"⚠️ 重点船舶","普通船舶")'
    ws[f"C{row}"].font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    ws[f"C{row}"].alignment = Alignment(horizontal="center", vertical="center")
    ws[f"C{row}"].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    ws[f"D{row}"] = f'=IF(C{row}="⚠️ 重点船舶","按重点船舶监管","")'
    ws[f"D{row}"].font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    ws[f"E{row}"] = "油轮/化学品/LNG/客船自动标记为重点船舶，需额外关注"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_keyship = row
    row += 1

    row += 1  # 空行

    # ========== 评估结果 ==========
    r_result_start = row
    ws.merge_cells(f"B{row}:E{row}")
    ws[f"B{row}"] = "八、评估结果"
    ws[f"B{row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    ws.row_dimensions[row].height = 24
    row += 1

    # 风险总分
    ws[f"B{row}"] = "风险总分（四项之和）"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    dim1_sum = f"SUM(D{r_age}:D{r_flag})"
    dim2_sum = f"SUM(D{r_psc}:D{r_dg})"
    dim3_sum = f"SUM(D{r_crew}:D{r_exp})"
    dim5_sum = f"SUM(D{r_draft}:D{r_cargo})"
    ws[f"C{row}"] = f"={dim1_sum}+{dim2_sum}+{dim3_sum}+{dim5_sum}"
    ws[f"C{row}"].style = "result_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=14, bold=True, color="C00000")
    ws[f"E{row}"] = "满分120分，得分越高风险越大"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_total = row
    row += 1

    # 一票否决触发
    ws[f"B{row}"] = "一票否决是否触发"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    veto_refs = ",".join([f"C{r}=\"是\"" for r in r_veto_items])
    ws[f"C{row}"] = f'=IF(OR({veto_refs}),"已触发","未触发")'
    ws[f"C{row}"].style = "result_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=12, bold=True)
    ws[f"E{row}"] = "任一票决项为\"是\"即触发"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_veto_flag = row
    row += 1

    # 风险等级
    ws[f"B{row}"] = "最终风险等级"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=12, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"C{row}:D{row}")
    ws[f"C{row}"] = f'=IF(C{r_veto_flag}="已触发","🔴 高风险",IF(C{r_total}<=25,"🟢 低风险",IF(C{r_total}<=50,"🟡 中风险","🔴 高风险")))'
    ws[f"C{row}"].style = "result_style"
    ws[f"C{row}"].font = Font(name="微软雅黑", size=16, bold=True)
    ws[f"E{row}"] = "🟢 ≤25分 | 🟡 26-50分 | 🔴 >50分 或一票否决"
    ws[f"E{row}"].font = Font(name="微软雅黑", size=9, color="666666")
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    r_level = row
    row += 1

    # 建议措施
    ws[f"B{row}"] = "建议措施"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"C{row}:E{row}")

    veto_msg = "一票否决触发，建议：①暂缓审批进港手续；②强制引航+拖轮协助；③登轮扩大检查范围；④必要时要求船方出具整改承诺"
    low_msg = "🟢 低风险：可予以通关便利，按常规比例抽查或免于登轮检查。"
    mid_msg = "🟡 中风险：建议安排登轮检查，重点核查本船历史缺陷项和船员配员情况。"
    high_msg = "🔴 高风险：建议①优先安排登轮检查；②扩大检查范围（消防安全+救生+航行安全）；③视情要求船方整改后作业；④重大风险可暂缓装卸货。"

    formula = (
        '=IF(C{vf}="已触发","{veto}",'
        'IF(C{lev}="🟢 低风险","{low}",'
        'IF(C{lev}="🟡 中风险","{mid}","{high}")))'
    ).format(vf=r_veto_flag, lev=r_level, veto=veto_msg, low=low_msg, mid=mid_msg, high=high_msg)
    ws[f"C{row}"] = formula
    ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f"C{row}"].font = Font(name="微软雅黑", size=10)
    ws.row_dimensions[row].height = 90
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # 重点风险关注清单（新增）
    ws[f"B{row}"] = "重点风险关注清单"
    ws[f"B{row}"].font = Font(name="微软雅黑", size=11, bold=True)
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(f"C{row}:E{row}")

    warn_parts = [
        f'IF(D{r_bridge}="❌ 禁止通航","【超高】该船可能超高，禁止通航；","")',
        f'IF(D{r_bridge}="⚠️ 高风险","【超高】桥梁净空不足，需特别关注；","")',
        f'IF(D{r_first}="⚠️ 高风险","【首航】船长及驾驶员均首次进港，建议强制引航；","")',
        f'IF(D{r_first}="△ 关注","【首航】船长首次进港，建议安排引航；","")',
        f'IF(D{r_berth}="⚠️ 高风险","【泊位】船舶明显超泊位设计吨级，需论证靠泊方案；","")',
        f'IF(C{r_keyship}="⚠️ 重点船舶","【重点船型】该船为油轮/化学品/液化气/客船，按重点船舶监管要求执行；","")',
        f'IF(C{r_pilot}="强制引航","【引航】强制引航船舶，确保引航员按时到位；","")',
    ]
    warn_formula = "=CONCATENATE(" + ",".join(warn_parts) + ")"
    ws[f"C{row}"] = warn_formula
    ws[f"C{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws[f"C{row}"].font = Font(name="微软雅黑", size=10, bold=True, color="C00000")
    ws.row_dimensions[row].height = 80
    for col in range(2, 6):
        ws.cell(row=row, column=col).border = thin_border
    row += 1

    # ========== 条件格式 ==========
    # 风险总分单元格根据数值变色
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    ws.conditional_formatting.add(f"C{r_total}",
        FormulaRule(formula=[f"$C${r_total}<=25"], fill=green_fill))
    ws.conditional_formatting.add(f"C{r_total}",
        FormulaRule(formula=[f"AND($C${r_total}>25,$C${r_total}<=50)"], fill=yellow_fill))
    ws.conditional_formatting.add(f"C{r_total}",
        FormulaRule(formula=[f"$C${r_total}>50"], fill=red_fill))

    # 风险等级单元格条件格式
    ws.conditional_formatting.add(f"C{r_level}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("低风险",$C${r_level}))'], fill=green_fill))
    ws.conditional_formatting.add(f"C{r_level}",
        FormulaRule(formula=[f'ISNUMBER(SEARCH("中风险",$C${r_level}))'], fill=yellow_fill))
    ws.conditional_formatting.add(f"C{r_level}",
        FormulaRule(formula=[f'OR(ISNUMBER(SEARCH("高风险",$C${r_level})),$C${r_veto_flag}="已触发")'], fill=red_fill))

    # ========== 风险雷达图 + 进度条（新增视觉增强）==========
    # 图表数据区（放在右侧 F-H 列，较窄，不干扰主表）
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8

    chart_base = 7
    ws[f"F{chart_base}"] = "维度"
    ws[f"G{chart_base}"] = "得分"
    ws[f"H{chart_base}"] = "满分"
    for c in ["F", "G", "H"]:
        ws[f"{c}{chart_base}"].font = Font(name="微软雅黑", size=9, bold=True, color="FFFFFF")
        ws[f"{c}{chart_base}"].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        ws[f"{c}{chart_base}"].alignment = Alignment(horizontal="center", vertical="center")

    radar_data = [
        ("船舶固有", f"=C{r_dim1_end}", 30),
        ("历史绩效", f"=C{r_dim2_end}", 40),
        ("船员配员", f"=C{r_dim3_end}", 25),
        ("航行通航", f"=C{r_dim5_end}", 25),
    ]
    for idx, (name, formula, max_val) in enumerate(radar_data, start=chart_base+1):
        ws[f"F{idx}"] = name
        ws[f"F{idx}"].font = Font(name="微软雅黑", size=9)
        ws[f"G{idx}"] = formula
        ws[f"G{idx}"].font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
        ws[f"G{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"H{idx}"] = max_val
        ws[f"H{idx}"].font = Font(name="微软雅黑", size=9, color="999999")
        ws[f"H{idx}"].alignment = Alignment(horizontal="center", vertical="center")

    # 创建雷达图
    chart = RadarChart()
    chart.type = "filled"
    chart.style = 10
    chart.title = "四维度风险雷达图"
    chart.y_axis.delete = True
    chart.width = 22
    chart.height = 14

    data_ref = Reference(ws, min_col=7, min_row=chart_base, max_row=chart_base+4)
    cats_ref = Reference(ws, min_col=6, min_row=chart_base+1, max_row=chart_base+4)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # 美化雷达图颜色
    from openpyxl.chart.series import DataPoint
    from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
    series = chart.series[0]
    series.graphicalProperties.solidFill = "4472C4"
    series.graphicalProperties.line.solidFill = "1F4E78"
    series.graphicalProperties.line.width = 20000  # 线宽

    ws.add_chart(chart, "F12")

    # 风险总分进度条（数据条）
    ws.conditional_formatting.add(f"C{r_total}",
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=120,
                    color="FF6C6C", showValue=True, minLength=None, maxLength=None))

    # 冻结窗格
    ws.freeze_panes = "B8"

    return ws


# ============ 创建评分标准参考表 ============
def create_std_sheet(wb, styles):
    ws = wb.create_sheet(title="评分标准速查")
    ws.sheet_properties.tabColor = "70AD47"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 15

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    row = 1
    ws.merge_cells("B1:D1")
    ws["B1"] = "评分标准速查表"
    ws["B1"].style = "title_style"
    for col in range(2, 5):
        ws.cell(row=1, column=col).border = thin_border
    row += 1

    standards = [
        ("船舶固有风险（满分30）", "", ""),
        ("船龄", "≤10年 / 10-15年 / 15-20年 / 20-25年 / >25年", "0 / 5 / 10 / 15 / 20"),
        ("船型", "普通货船/集装箱 / 散货船 / 油轮/化学品/LNG / 客船/滚装", "0 / 3 / 5 / 4"),
        ("船旗国", "白名单 / 灰名单 / 黑名单 / 未知", "0 / 5 / 10 / 3"),
        ("", "", ""),
        ("历史安全绩效（满分40）", "", ""),
        ("近12月PSC检查", "0次 / 1次 / 2次 / ≥3次", "0 / 3 / 8 / 15"),
        ("近36月滞留", "0次 / 1次 / 2次 / ≥3次", "0 / 10 / 20 / 30"),
        ("近12月缺陷数", "0个 / 1-3个 / 4-6个 / ≥7个", "0 / 5 / 10 / 15"),
        ("东莞历史记录", "无记录 / 有缺陷 / 曾滞留", "0 / 5 / 15"),
        ("", "", ""),
        ("船员与配员风险（满分25）", "", ""),
        ("配员符合率", "100% / 90-99% / 80-89% / 70-79% / <70%", "0 / 5 / 10 / 15 / 25"),
        ("证书过期", "无过期 / 1-2人 / ≥3人", "0 / 5 / 12"),
        ("关键岗位资历", "≥12月 / 6-12月 / 3-6月 / <3月 / 新上船", "0 / 2 / 5 / 8 / 10"),
        ("", "", ""),
        ("航行与通航风险（满分25）", "", ""),
        ("吃水余量", "≥2m / 1-2m / 0.5-1m / 0-0.5m / <0m", "0 / 5 / 10 / 20 / 25"),
        ("到港时段", "白天 / 夜间 / 凌晨(00-06)", "0 / 3 / 5"),
        ("气象条件", "良好 / 一般 / 恶劣", "0 / 5 / 12"),
        ("载货类型", "普通货 / 包装危险品 / 散装油/化学品", "0 / 5 / 10"),
        ("", "", ""),
        ("风险等级划分", "", ""),
        ("🟢 低风险", "总分 ≤ 25分", "常规抽查或免查"),
        ("🟡 中风险", "总分 26-50分", "安排登轮检查，重点核查"),
        ("🔴 高风险", "总分 > 50分 或 一票否决触发", "优先检查+扩大范围+视情整改"),
        ("", "", ""),
        ("一票否决项", "", ""),
        ("1", "近6个月内有PSC滞留记录", "直接高风险"),
        ("2", "船龄>25年且为油轮/化学品船/液化气船", "直接高风险"),
        ("3", "配员符合率<70%", "直接高风险"),
        ("4", "吃水余量<0米（超吃水）", "直接高风险"),
        ("5", "关键船员证书过期≥3人或船长证书过期", "直接高风险"),
        ("", "", ""),
        ("重点风险提示（不纳入总分）", "", ""),
        ("桥梁净空", "≥3m / 1-3m / 0-1m / <0m", "正常/关注/⚠️/禁止通航"),
        ("驾驶员首航", "否 / 船长首次 / 均首次", "熟悉/△关注/⚠️高风险"),
        ("引航安排", "无需 / 申请 / 强制", "自航/有引航/强制引航"),
        ("泊位匹配", "匹配 / 略超 / 明显超配", "正常/关注/⚠️高风险"),
        ("重点船舶", "油轮/化学品/LNG/客船自动标记", "需重点监管"),
    ]

    for item in standards:
        label, desc, score = item
        ws[f"B{row}"] = label
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = score
        if label.endswith(")") and "满分" in label:
            ws[f"B{row}"].style = "dim_style"
            ws[f"C{row}"].style = "dim_style"
            ws[f"D{row}"].style = "dim_style"
            ws.row_dimensions[row].height = 22
        elif label in ("风险等级划分", "一票否决项"):
            ws[f"B{row}"].style = "dim_style"
            ws[f"C{row}"].style = "dim_style"
            ws[f"D{row}"].style = "dim_style"
            ws.row_dimensions[row].height = 22
        elif label in ("🟢 低风险", "🟡 中风险", "🔴 高风险"):
            ws[f"B{row}"].font = Font(name="微软雅黑", size=10, bold=True)
            ws[f"C{row}"].font = Font(name="微软雅黑", size=10)
            ws[f"D{row}"].font = Font(name="微软雅黑", size=10)
        else:
            ws[f"B{row}"].style = "label_style"
            ws[f"C{row}"].font = Font(name="微软雅黑", size=9, color="444444")
            ws[f"D{row}"].font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
        for col in range(2, 5):
            ws.cell(row=row, column=col).border = thin_border
        row += 1

    ws.freeze_panes = "B3"
    return ws


# ============ 创建首靠船日报清单 ============
def create_daily_sheet(wb, styles):
    ws = wb.create_sheet(title="日报清单", index=1)
    ws.sheet_properties.tabColor = "ED7D31"

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 14
    ws.column_dimensions["M"].width = 12
    ws.column_dimensions["N"].width = 10

    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # 标题
    ws.merge_cells("B1:N1")
    ws["B1"] = "首靠船日报清单（东莞辖区）"
    ws["B1"].style = "title_style"
    for col in range(2, 15):
        ws.cell(row=1, column=col).border = thin_border
    ws.row_dimensions[1].height = 32

    # 日期行
    ws.merge_cells("B2:N2")
    ws["B2"] = '日期：______________    值班长签字：______________    共 ____ 艘首靠船'
    ws["B2"].font = Font(name="微软雅黑", size=11, bold=True)
    ws["B2"].alignment = Alignment(horizontal="left", vertical="center")
    for col in range(2, 15):
        ws.cell(row=2, column=col).border = thin_border
    ws.row_dimensions[2].height = 26

    # 表头
    headers = [
        "序号", "日期", "船名", "IMO", "ETA", "船型",
        "风险等级", "重点船舶", "桥梁净空", "驾驶员首航",
        "引航", "泊位匹配", "一票否决", "评估人"
    ]
    for col_idx, header in enumerate(headers, start=2):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.style = "header_style"
        cell.border = thin_border
    ws.row_dimensions[3].height = 28

    # 预填示例数据（供用户参考格式）
    examples = [
        ["1", "2026-05-10", "海发轮", "9876543", "08:30", "散货船", "🟡 中风险", "普通船舶", "✓ 正常", "✓ 熟悉", "自航", "✓ 正常", "未触发", "张三"],
        ["2", "2026-05-10", "东方油1", "8765432", "14:00", "油轮", "🔴 高风险", "⚠️ 重点船舶", "✓ 正常", "△ 关注", "强制引航", "✓ 正常", "未触发", "李四"],
        ["3", "2026-05-10", "长江之星", "7654321", "22:30", "集装箱船", "🟢 低风险", "普通船舶", "✓ 正常", "✓ 熟悉", "申请引航", "✓ 正常", "未触发", "王五"],
    ]

    for r_idx, data in enumerate(examples, start=4):
        for c_idx, val in enumerate(data, start=2):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(name="微软雅黑", size=9)
            # 风险等级列条件着色
            if c_idx == 8:  # G列 = 风险等级
                if "🟢" in str(val):
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif "🟡" in str(val):
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif "🔴" in str(val):
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            # 重点船舶列着色
            if c_idx == 9 and "⚠️" in str(val):  # H列 = 重点船舶
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(name="微软雅黑", size=9, bold=True, color="C00000")
        ws.row_dimensions[r_idx].height = 22

    # 空行（留给用户填写）+ 条件格式预置
    for empty_row in range(7, 27):
        for col in range(2, 15):
            ws.cell(row=empty_row, column=col).border = thin_border
        ws.row_dimensions[empty_row].height = 22

    # 预置日报清单的条件格式（用户填写后自动生效）
    # 一票否决列：已触发标红
    red_text = Font(name="微软雅黑", size=9, bold=True, color="C00000")
    for r in range(7, 27):
        cell = ws.cell(row=r, column=13)  # M列 = 一票否决
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=9)
    # 桥梁净空列：含❌或⚠️标红
    for r in range(7, 27):
        cell = ws.cell(row=r, column=9)  # I列 = 桥梁净空
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="微软雅黑", size=9)

    # 当日统计区
    stat_row = 28
    ws.merge_cells(f"B{stat_row}:E{stat_row}")
    ws[f"B{stat_row}"] = "当日统计"
    ws[f"B{stat_row}"].style = "dim_style"
    for col in range(2, 6):
        ws.cell(row=stat_row, column=col).border = thin_border
    ws.row_dimensions[stat_row].height = 24

    stats = [
        ("🟢 低风险", "=COUNTIF(G4:G26,\"*低风险*\")", "C6EFCE"),
        ("🟡 中风险", "=COUNTIF(G4:G26,\"*中风险*\")", "FFEB9C"),
        ("🔴 高风险", "=COUNTIF(G4:G26,\"*高风险*\")", "FFC7CE"),
        ("⚠️ 重点船舶", "=COUNTIF(H4:H26,\"*重点船舶*\")", "FFC7CE"),
        ("❌ 一票否决触发", "=COUNTIF(M4:M26,\"已触发\")", "FFC7CE"),
    ]
    for idx, (label, formula, color) in enumerate(stats, start=stat_row+1):
        ws[f"B{idx}"] = label
        ws[f"B{idx}"].font = Font(name="微软雅黑", size=10, bold=True)
        ws[f"B{idx}"].alignment = Alignment(horizontal="left", vertical="center")
        ws[f"C{idx}"] = formula
        ws[f"C{idx}"].font = Font(name="微软雅黑", size=11, bold=True, color="C00000")
        ws[f"C{idx}"].alignment = Alignment(horizontal="center", vertical="center")
        ws[f"C{idx}"].fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.merge_cells(f"D{idx}:E{idx}")
        ws[f"D{idx}"] = "艘"
        ws[f"D{idx}"].font = Font(name="微软雅黑", size=10)
        ws[f"D{idx}"].alignment = Alignment(horizontal="left", vertical="center")
        for col in range(2, 6):
            ws.cell(row=idx, column=col).border = thin_border
        ws.row_dimensions[idx].height = 24

    # 底部说明
    note_row = stat_row + len(stats) + 1
    ws.merge_cells(f"B{note_row}:N{note_row}")
    ws[f"B{note_row}"] = "说明：每日将当天所有首靠船的关键信息填入本表，实现'一眼掌握'当日风险态势。统计区会自动汇总各等级数量。风险等级和重点标记请从单船评估表中复制。"
    ws[f"B{note_row}"].font = Font(name="微软雅黑", size=9, italic=True, color="666666")
    ws[f"B{note_row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 36

    # 冻结窗格
    ws.freeze_panes = "B4"

    return ws


# ============ 主程序 ============
def main():
    wb = openpyxl.Workbook()
    styles = create_styles(wb)

    # 创建评估表
    create_eval_sheet(wb, styles)

    # 创建日报清单
    create_daily_sheet(wb, styles)

    # 创建评分标准速查表
    create_std_sheet(wb, styles)

    # 保存
    output_path = "首靠船风险评估表.xlsx"
    wb.save(output_path)
    print(f"✅ 已成功生成：{output_path}")
    print("\n使用说明：")
    print("1. 打开 Excel 文件，在\"评估表\"中依次选择各项目的下拉选项")
    print("2. 分值和风险等级会自动计算")
    print("3. 一票否决项如有触发，最终等级强制为🔴高风险")
    print("4. 参考\"评分标准速查\"表了解各指标对应分值")
    print("5. 使用\"日报清单\"汇总当日所有首靠船，便于快速掌握整体态势")


if __name__ == "__main__":
    main()
