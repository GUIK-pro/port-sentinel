#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
进出港计划上传模板生成脚本（东莞辖区）
运行后生成：进出港计划上传模板.xlsx

设计依据：《首靠船风险识别工作指引》V2.2 §4.2 上传字段
- 基本信息（4） · 船舶属性（6） · 行程时点（6） · 船舶联络（2） · 通航安全（4）
- 共 22 列，全部必填（备注列允许空）
"""

import sys

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("正在安装 openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "-q"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation


# ============ 设计 Token ============
COLOR = {
    "naval":   "0A1F44",
    "slate":   "475569",
    "teal":    "00A896",
    "amber":   "D97706",
    "red":     "DC2626",
    "ink":     "0B1220",
    "ink_2":   "475569",
    "line":    "E6E9EF",
    "bg":      "F8FAFC",
    "warn_bg": "FEF3C7",
    "ok_bg":   "ECFDF5",
}

FONT_HAN = "微软雅黑"
FONT_NUM = "Consolas"

# ============ 字段定义（5 组共 22 列） ============
# 字段类型说明：
#   SEQ=序号 / DATE=日期 / TEXT=文本 / NUM=小数 / INT=整数 / PHONE=11位手机
#   BERTH=泊位下拉 / SHIP=船型下拉 / CARGO=货种下拉 / YN=是否下拉 / DT_TXT=时间文本

FIELD_GROUPS = [
    {
        "key": "BASIC", "label": "基本信息", "color": COLOR["naval"],
        "fields": [
            ("序号",       "顺序号",                    "SEQ"),
            ("日期",       "M.D 或 YYYY-MM-DD",         "DATE"),
            ("靠泊码头",   "下拉选择",                  "BERTH"),
            ("船名",       "中文船名",                  "TEXT"),
        ],
    },
    {
        "key": "SHIP", "label": "船舶属性", "color": COLOR["slate"],
        "fields": [
            ("船舶类型",       "下拉选择",            "SHIP"),
            ("船舶总吨",       "GT 数值",             "INT"),
            ("船舶最大长度",   "米",                  "NUM"),
            ("甲板层数",       "1-5 层",              "INT_S"),
            ("货物种类",       "下拉选择",            "CARGO"),
            ("货物数量",       "吨/标箱",             "INT"),
        ],
    },
    {
        "key": "VOYAGE", "label": "行程时点", "color": COLOR["teal"],
        "fields": [
            ("进港吃水",                "米",                              "NUM"),
            ("出港吃水",                "米",                              "NUM"),
            ("进港水面以上最大高度",    "米",                              "NUM"),
            ("出港水面以上最大高度",    "米",                              "NUM"),
            ("进港时间",                "如：3月13日1200时",               "DT_TXT"),
            ("离港时间",                "如：3月13日1700时",               "DT_TXT"),
        ],
    },
    {
        "key": "CONTACT", "label": "船舶联络", "color": COLOR["amber"],
        "fields": [
            ("船长姓名",   "当前在任船长",     "TEXT"),
            ("联系方式",   "11 位手机号",      "PHONE"),
        ],
    },
    {
        "key": "SAFETY", "label": "通航安全", "color": COLOR["red"],
        "fields": [
            ("航经水道及最低桥梁",                                "如：太平航道·虎门大桥(60m)",  "TEXT"),
            ("是否通知船舶需要经过桥梁的净空高度",                "是 / 否 / 不适用",              "YN"),
            ("是否通知船舶检查主机电机等通信设备",                "是 / 否 / 不适用",              "YN"),
            ("备注",                                              "可空",                          "TEXT"),
        ],
    },
]

# 下拉枚举
BERTHS = [
    "中泰", "富之源", "石角", "立沙岛", "新沙",
    "麻涌", "沙田", "虎门", "长安", "锚地", "其他",
]
SHIP_TYPES = [
    "散货船", "件杂货船", "集装箱船", "油船", "化学品船",
    "液化气船", "客船", "汽车滚装", "工程船", "其他",
]
CARGO_TYPES = [
    "石子", "碎石", "玉米", "煤炭", "矿石", "粮食",
    "钢材", "原油", "成品油", "化学品", "液化气",
    "集装箱", "件杂货", "其他",
]
YN_OPTS = ["是", "否", "不适用"]

# 示例数据（来自用户实际样例，保持原样）
SAMPLE_ROWS = [
    [1, "3.13", "中泰",    "粤韶关货2200",  "散货船", 495, 49, "石子", 1000,
     "3.2", "4.5", "4", "7",  "3月13日1200时", "3月13日1700时",
     "韶关货",   "12312312344", "太平航道·虎门大桥(60m)", "是", "是", ""],
    [2, "3.13", "中泰",    "桂桂平货6989",  "散货船", 595, 50, "碎石", 1100,
     "4.2", "5.8", "4", "6",  "3月13日1200时", "3月13日1700时",
     "张三",     "13924189237", "太平航道·虎门大桥(60m)", "是", "是", ""],
    [3, "3.13", "富之源",  "益嘉26",        "散货船", 678, 50, "玉米", 1100,
     "4.2", "5.8", "4", "6",  "3月13日1200时", "3月13日1700时",
     "李四",     "13924189238", "蕉门水道·南沙大桥(38m)", "是", "是", ""],
    [4, "3.13", "石角",    "桂平宏远838",   "散货船", 678, 50, "玉米", 1100,
     "4.2", "5.8", "4", "6",  "3月14日0600时", "3月14日1200时",
     "李四",     "13924189238", "蕉门水道·南沙大桥(38m)", "是", "是", ""],
    [5, "3.13", "中泰",    "桂桂平货6989",  "散货船", 595, 50, "碎石", 1100,
     "4.2", "5.8", "4", "6",  "3月14日1200时", "3月14日1700时",
     "韶关货",   "12312312344", "太平航道·虎门大桥(60m)", "是", "是", ""],
]


# ============ 通用样式 ============
def thin_border():
    side = Side(style="thin", color=COLOR["line"])
    return Border(left=side, right=side, top=side, bottom=side)


def fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


# ============ Sheet 1：进出港计划 ============
def build_main_sheet(wb):
    ws = wb.active
    ws.title = "进出港计划"

    total_cols = sum(len(g["fields"]) for g in FIELD_GROUPS)

    # 第 1 行：标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(row=1, column=1, value="东莞海事局 · 进出港计划上传模板（V2.2 内河水网字段适配版）")
    t.font = Font(name=FONT_HAN, size=15, bold=True, color="FFFFFF")
    t.fill = fill(COLOR["naval"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 第 2 行：上传规则提示
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    tip = ws.cell(row=2, column=1, value=(
        "上传规则：① 单文件 ≤ 5MB · 仅支持 .xlsx/.xls   "
        "② 22 列字段（备注可空）· 行号 1-3 为表头不可删   "
        "③ 上传成功 = 写入近 6 个月滚动库 · 触发首靠比对（船名 + 总吨 + 船长 + 电话）   "
        "④ 桥梁净空 / 主机电机检查通知 必须如实标记"
    ))
    tip.font = Font(name=FONT_HAN, size=10, color=COLOR["ink_2"])
    tip.fill = fill(COLOR["warn_bg"])
    tip.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
    ws.row_dimensions[2].height = 38

    # 第 3 行：分组表头
    col = 1
    for g in FIELD_GROUPS:
        span = len(g["fields"])
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + span - 1)
        c = ws.cell(row=3, column=col, value=f"{g['label']} · {g['key']}")
        c.font = Font(name=FONT_HAN, size=11, bold=True, color="FFFFFF")
        c.fill = fill(g["color"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
        col += span
    ws.row_dimensions[3].height = 28

    # 第 4 行：字段名
    col = 1
    field_meta = []
    for g in FIELD_GROUPS:
        for fname, hint, ftype in g["fields"]:
            star = "" if fname == "备注" else "  *"
            cell = ws.cell(row=4, column=col, value=f"{fname}{star}")
            cell.font = Font(name=FONT_HAN, size=10, bold=True, color=COLOR["ink"])
            cell.fill = fill(COLOR["bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
            field_meta.append((col, fname, hint, ftype))
            col += 1
    ws.row_dimensions[4].height = 36

    # 第 5 行：提示行
    for col_, fname, hint, ftype in field_meta:
        c = ws.cell(row=5, column=col_, value=hint)
        c.font = Font(name=FONT_HAN, size=9, italic=True, color=COLOR["ink_2"])
        c.fill = fill("F1F5F9")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[5].height = 26

    # 第 6 行起：示例数据
    sample_start = 6
    for r_idx, row_data in enumerate(SAMPLE_ROWS):
        r = sample_start + r_idx
        for col_, value in enumerate(row_data, start=1):
            ftype = field_meta[col_ - 1][3]
            font_name = FONT_NUM if ftype in ("SEQ", "INT", "INT_S", "NUM", "PHONE") else FONT_HAN
            cell = ws.cell(row=r, column=col_, value=value)
            cell.font = Font(name=font_name, size=10, color=COLOR["ink"])
            cell.fill = fill(COLOR["ok_bg"])
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()
        ws.row_dimensions[r].height = 22

    sample_end = sample_start + len(SAMPLE_ROWS) - 1
    note_row = sample_end + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=total_cols)
    note = ws.cell(row=note_row, column=1,
                   value=f"↑ 第 {sample_start}-{sample_end} 行为示例数据（绿底），请清空后填写当日实际计划")
    note.font = Font(name=FONT_HAN, size=9, italic=True, color=COLOR["amber"])
    note.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[note_row].height = 22

    # 列宽
    width_map = {
        "SEQ": 6, "DATE": 12, "BERTH": 12, "TEXT": 18, "INT": 11,
        "INT_S": 10, "NUM": 11, "DT_TXT": 18, "SHIP": 13, "CARGO": 12,
        "PHONE": 14, "YN": 10,
    }
    for col_, fname, hint, ftype in field_meta:
        base = width_map.get(ftype, 14)
        ws.column_dimensions[get_column_letter(col_)].width = max(base, len(fname) + 3)

    # 数据验证
    data_end = sample_start + 200

    def add_list_dv(col_, options, title, msg):
        dv = DataValidation(
            type="list",
            formula1='"' + ",".join(options) + '"',
            allow_blank=False, showErrorMessage=True,
        )
        dv.errorTitle = title
        dv.error = msg
        ws.add_data_validation(dv)
        letter = get_column_letter(col_)
        dv.add(f"{letter}{sample_start}:{letter}{data_end}")

    def add_int_dv(col_, lo, hi, title, msg):
        dv = DataValidation(
            type="whole", operator="between",
            formula1=str(lo), formula2=str(hi),
            allow_blank=False, showErrorMessage=True,
        )
        dv.errorTitle = title; dv.error = msg
        ws.add_data_validation(dv)
        letter = get_column_letter(col_)
        dv.add(f"{letter}{sample_start}:{letter}{data_end}")

    def add_phone_dv(col_):
        dv = DataValidation(
            type="textLength", operator="equal", formula1="11",
            allow_blank=False, showErrorMessage=True,
        )
        dv.errorTitle = "电话格式异常"
        dv.error = "联系方式须为 11 位手机号"
        ws.add_data_validation(dv)
        letter = get_column_letter(col_)
        dv.add(f"{letter}{sample_start}:{letter}{data_end}")

    for col_, fname, hint, ftype in field_meta:
        if ftype == "BERTH":
            add_list_dv(col_, BERTHS, "码头无效", "请从下拉列表选择辖区合法码头")
        elif ftype == "SHIP":
            add_list_dv(col_, SHIP_TYPES, "船型无效", "请从下拉列表选择标准船型")
        elif ftype == "CARGO":
            add_list_dv(col_, CARGO_TYPES, "货种无效", "请从下拉列表选择标准货种")
        elif ftype == "YN":
            add_list_dv(col_, YN_OPTS, "选项无效", "请选择 是 / 否 / 不适用")
        elif ftype == "PHONE":
            add_phone_dv(col_)
        elif ftype == "INT":
            add_int_dv(col_, 0, 500000, "数值异常", "请输入 0-500000 整数")
        elif ftype == "INT_S":
            add_int_dv(col_, 1, 10, "层数异常", "甲板层数 1-10 之间")
        elif ftype == "SEQ":
            add_int_dv(col_, 1, 9999, "序号异常", "序号 1-9999")

    ws.freeze_panes = "B6"
    return ws


# ============ Sheet 2：填写说明 ============
def build_help_sheet(wb):
    ws = wb.create_sheet("填写说明")

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="填写说明 · 字段释义 · 入库规则（V2.2）")
    t.font = Font(name=FONT_HAN, size=14, bold=True, color="FFFFFF")
    t.fill = fill(COLOR["naval"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = ["字段类别", "字段名", "必填", "说明 / 格式要求"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(name=FONT_HAN, size=10, bold=True, color="FFFFFF")
        c.fill = fill(COLOR["slate"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 24

    row = 3
    for g in FIELD_GROUPS:
        for fname, hint, ftype in g["fields"]:
            req = "—" if fname == "备注" else "✅"
            cells = [
                (1, g["label"], g["color"], "FFFFFF", True,  False),
                (2, fname,      None,        None,     True,  False),
                (3, req,        None,        None,     False, False),
                (4, hint,       None,        None,     False, True),
            ]
            for col_, val, bg, fg, bold, wrap in cells:
                c = ws.cell(row=row, column=col_, value=val)
                c.font = Font(name=FONT_HAN, size=10, bold=bold, color=fg or COLOR["ink"])
                if bg:
                    c.fill = fill(bg)
                c.alignment = Alignment(
                    horizontal="left" if col_ == 4 else "center",
                    vertical="center", wrap_text=wrap, indent=1 if col_ == 4 else 0,
                )
                c.border = thin_border()
            ws.row_dimensions[row].height = 22
            row += 1

    # 入库规则
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    c = ws.cell(row=row, column=1, value="入库与首靠比对规则（V2.2）")
    c.font = Font(name=FONT_HAN, size=12, bold=True, color="FFFFFF")
    c.fill = fill(COLOR["teal"])
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 28
    row += 1

    rules = [
        "① 上传即入库：校验通过后秒级写入「近 6 个月进出港计划库」，自动触发首靠比对",
        "② 比对字段：船名 + 船舶总吨 + 船长姓名 + 联系方式（任一未匹配即视为首靠）",
        "③ 同船覆盖：同船多次报送以最新一次为准（保留时间戳供审计）",
        "④ 滚动 6 月：每日 0 点自动归档 > 6 月数据至历史库",
        "⑤ 失败队列：异常行进入待处理队列，需修正后重传",
        "⑥ 安全确认必填：桥梁净空通知 / 主机电机检查通知 两列必须如实标记，是法定安全告知动作",
        "⑦ 上传时点：建议次日进港船舶在到港前 24 小时（T-24h）完成上传",
    ]
    for r in rules:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        c = ws.cell(row=row, column=1, value=r)
        c.font = Font(name=FONT_HAN, size=10, color=COLOR["ink"])
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        c.border = thin_border()
        ws.row_dimensions[row].height = 32
        row += 1

    widths = {1: 14, 2: 28, 3: 8, 4: 56}
    for col_, w in widths.items():
        ws.column_dimensions[get_column_letter(col_)].width = w


# ============ Sheet 3：参考枚举 ============
def build_ref_sheet(wb):
    ws = wb.create_sheet("参考枚举")

    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="参考枚举 · 码头 / 船型 / 货种 / 是否")
    t.font = Font(name=FONT_HAN, size=13, bold=True, color="FFFFFF")
    t.fill = fill(COLOR["naval"])
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    columns = [
        (1, "靠泊码头",   COLOR["naval"], BERTHS),
        (2, "船舶类型",   COLOR["slate"], SHIP_TYPES),
        (3, "货物种类",   COLOR["amber"], CARGO_TYPES),
        (4, "是否选项",   COLOR["red"],   YN_OPTS),
    ]
    for col_, label, color, _ in columns:
        c = ws.cell(row=2, column=col_, value=label)
        c.font = Font(name=FONT_HAN, size=11, bold=True, color="FFFFFF")
        c.fill = fill(color)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[2].height = 24

    max_len = max(len(lst) for _, _, _, lst in columns)
    for i in range(max_len):
        for col_, _, _, lst in columns:
            val = lst[i] if i < len(lst) else ""
            c = ws.cell(row=3 + i, column=col_, value=val)
            c.font = Font(name=FONT_HAN, size=10, color=COLOR["ink"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()

    for col_, w in {1: 14, 2: 16, 3: 16, 4: 12}.items():
        ws.column_dimensions[get_column_letter(col_)].width = w


# ============ 主流程 ============
def main():
    wb = openpyxl.Workbook()
    build_main_sheet(wb)
    build_help_sheet(wb)
    build_ref_sheet(wb)

    out_path = "进出港计划上传模板.xlsx"
    wb.save(out_path)
    print(f"✅ 已生成：{out_path}")
    print(f"   · 主表：22 列字段（5 组）· 含 5 行示例 · 数据验证已启用")
    print(f"   · 副表：填写说明 + 参考枚举（码头/船型/货种/是否）")
    print(f"   · 表头冻结：前 5 行 + A 列")


if __name__ == "__main__":
    main()
